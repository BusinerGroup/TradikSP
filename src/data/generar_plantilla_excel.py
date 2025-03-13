#!/usr/bin/env python3
"""
Script para generar una plantilla Excel para el registro de clientes de TradikSP.
Este script crea un archivo Excel con validaciones de datos y ejemplos.
"""

import csv
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: Se requiere la biblioteca openpyxl.")
    print("Instálela con: pip install openpyxl")
    sys.exit(1)

def crear_plantilla_excel():
    # Definir la ruta del archivo CSV y Excel
    directorio_actual = os.path.dirname(os.path.abspath(__file__))
    archivo_csv = os.path.join(directorio_actual, 'plantilla_clientes.csv')
    archivo_excel = os.path.join(directorio_actual, 'plantilla_clientes.xlsx')
    
    # Verificar si existe el archivo CSV
    if not os.path.exists(archivo_csv):
        print(f"Error: No se encontró el archivo {archivo_csv}")
        return
    
    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Leer el archivo CSV
    with open(archivo_csv, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    centered_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Aplicar estilos a los encabezados
    for col_idx in range(1, 11):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment
        cell.border = border
    
    # Ajustar el ancho de las columnas
    column_widths = {
        1: 8,    # id
        2: 30,   # nombre
        3: 20,   # tipoDocumento
        4: 20,   # numeroDocumento
        5: 40,   # direccion
        6: 20,   # ciudad
        7: 25,   # contacto
        8: 15,   # telefono
        9: 30,   # email
        10: 25,  # tipoCliente
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    
    # Agregar validaciones de datos
    # Validación para tipoDocumento
    tipo_documento_dv = DataValidation(
        type="list",
        formula1='"Cédula de Ciudadanía,NIT,Pasaporte"',
        allow_blank=False
    )
    tipo_documento_dv.error = "Por favor, seleccione un tipo de documento válido."
    tipo_documento_dv.errorTitle = "Tipo de documento inválido"
    tipo_documento_dv.prompt = "Seleccione un tipo de documento"
    tipo_documento_dv.promptTitle = "Tipo de documento"
    ws.add_data_validation(tipo_documento_dv)
    tipo_documento_dv.add(f"C2:C1000")
    
    # Validación para tipoCliente
    tipo_cliente_dv = DataValidation(
        type="list",
        formula1='"Venue,OPC,Productor,Centro de Convenciones,Hotel,Restaurante,Wedding & Event Planner,Colega,Agencia de Publicidad,Institucional"',
        allow_blank=False
    )
    tipo_cliente_dv.error = "Por favor, seleccione un tipo de cliente válido."
    tipo_cliente_dv.errorTitle = "Tipo de cliente inválido"
    tipo_cliente_dv.prompt = "Seleccione un tipo de cliente"
    tipo_cliente_dv.promptTitle = "Tipo de cliente"
    ws.add_data_validation(tipo_cliente_dv)
    tipo_cliente_dv.add(f"J2:J1000")
    
    # Validación para ciudad
    ciudad_dv = DataValidation(
        type="list",
        formula1='"Bogotá,Medellín,Cali,Barranquilla,Cartagena,Bucaramanga,Pereira,Santa Marta,Manizales,Villavicencio,Pasto,Otra"',
        allow_blank=False
    )
    ciudad_dv.error = "Por favor, seleccione una ciudad válida."
    ciudad_dv.errorTitle = "Ciudad inválida"
    ciudad_dv.prompt = "Seleccione una ciudad"
    ciudad_dv.promptTitle = "Ciudad"
    ws.add_data_validation(ciudad_dv)
    ciudad_dv.add(f"F2:F1000")
    
    # Validación para numeroDocumento (solo números)
    numero_documento_dv = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False
    )
    numero_documento_dv.error = "Por favor, ingrese solo números."
    numero_documento_dv.errorTitle = "Número de documento inválido"
    numero_documento_dv.prompt = "Ingrese solo dígitos numéricos"
    numero_documento_dv.promptTitle = "Número de documento"
    ws.add_data_validation(numero_documento_dv)
    numero_documento_dv.add(f"D2:D1000")
    
    # Validación para telefono (solo números)
    telefono_dv = DataValidation(
        type="whole",
        operator="greaterThan",
        formula1="0",
        allow_blank=False
    )
    telefono_dv.error = "Por favor, ingrese solo números."
    telefono_dv.errorTitle = "Número de teléfono inválido"
    telefono_dv.prompt = "Ingrese solo dígitos numéricos"
    telefono_dv.promptTitle = "Teléfono"
    ws.add_data_validation(telefono_dv)
    telefono_dv.add(f"H2:H1000")
    
    # Agregar una hoja de instrucciones
    ws_instrucciones = wb.create_sheet(title="Instrucciones")
    
    instrucciones = [
        ["INSTRUCCIONES PARA COMPLETAR LA PLANTILLA DE CLIENTES"],
        [""],
        ["1. Complete los datos de cada cliente en una fila separada en la hoja 'Clientes'."],
        ["2. Respete los valores permitidos para los campos que tienen listas desplegables."],
        ["3. Asegúrese de que el campo 'id' sea único para cada cliente."],
        ["4. Para los campos numéricos (numeroDocumento, telefono), ingrese solo dígitos sin espacios, guiones u otros caracteres."],
        ["5. No modifique los encabezados de las columnas."],
        [""],
        ["DESCRIPCIÓN DE LOS CAMPOS:"],
        [""],
        ["id", "Identificador único del cliente. Debe ser un número entero positivo."],
        ["nombre", "Nombre completo del cliente o empresa."],
        ["tipoDocumento", "Tipo de documento de identidad. Seleccione de la lista desplegable: Cédula de Ciudadanía, NIT o Pasaporte."],
        ["numeroDocumento", "Número del documento de identidad. Solo dígitos numéricos."],
        ["direccion", "Dirección completa del cliente."],
        ["ciudad", "Ciudad donde se encuentra el cliente. Seleccione de la lista desplegable."],
        ["contacto", "Nombre de la persona de contacto."],
        ["telefono", "Teléfono de contacto. Solo dígitos numéricos."],
        ["email", "Correo electrónico. Debe tener un formato válido."],
        ["tipoCliente", "Categoría del cliente. Seleccione de la lista desplegable: Venue, OPC, Productor, Centro de Convenciones, Hotel, Restaurante, Wedding & Event Planner, Colega, Agencia de Publicidad o Institucional."],
        [""],
        ["NOTAS IMPORTANTES:"],
        [""],
        ["- El archivo incluye ejemplos de diferentes tipos de clientes que pueden usarse como referencia, con datos adaptados al contexto colombiano."],
        ["- Los números de teléfono deben ingresarse sin espacios, guiones o paréntesis."],
        ["- Los números de documento deben contener solo dígitos."],
        ["- Una vez completada la plantilla, guarde el archivo y proporcióneselo al administrador del sistema."],
        ["- Para cualquier duda o aclaración, contacte al soporte técnico."],
    ]
    
    for row_idx, row_data in enumerate(instrucciones, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_instrucciones.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif len(row_data) == 1 and row_data[0].startswith("DESCRIPCIÓN") or row_data[0].startswith("NOTAS"):
                cell.font = Font(bold=True, size=12)
            elif len(row_data) == 2 and row_idx >= 11 and row_idx <= 20:
                if col_idx == 1:
                    cell.font = Font(bold=True)
    
    # Ajustar el ancho de las columnas en la hoja de instrucciones
    ws_instrucciones.column_dimensions['A'].width = 20
    ws_instrucciones.column_dimensions['B'].width = 80
    
    # Guardar el archivo Excel
    wb.save(archivo_excel)
    print(f"Plantilla Excel creada exitosamente: {archivo_excel}")

if __name__ == "__main__":
    crear_plantilla_excel() 